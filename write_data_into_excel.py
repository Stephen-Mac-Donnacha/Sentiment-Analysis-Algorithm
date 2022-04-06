import openpyxl

from main import return_profile_data

# Workbook saved here as a global variable
wb = openpyxl.Workbook()


def get_profile_data():
    data_arr = return_profile_data()

    for i in range(len(data_arr)):
        data1 = data_arr[0]
        data2 = data_arr[1]

    return data1, data2


# Store the sentiment and frequency data here
sentiment_data = get_profile_data()[0]
frequency_data = get_profile_data()[1]


# Method to write sentiment data to excel
def write_sentiment_data_to_excel():
    # Create worksheet for sentiment data
    sentiment_sheet = wb.create_sheet("Sentiment Data")

    sentiment_sheet['A1'] = "Date"
    sentiment_sheet['B1'] = "Sentiment"

    # For each entry, write it into the excel spreadsheet
    for row, (Date, Sentiment) in enumerate(sentiment_data.items(), start=2):
        sentiment_sheet[f"A{row}"] = Date
        sentiment_sheet[f"B{row}"] = Sentiment

    wb.save("C:\\Users\\Stephen Mac Donnacha\\PycharmProjects\\Final Year Project\\excel-worksheets\\fyp-data.xlsx")


# Method to write frequency data to excel
def write_frequency_data_to_excel():
    # Create worksheet for frequency data
    frequency_sheet = wb.create_sheet("Frequency Data")

    frequency_sheet['A1'] = "Date"
    frequency_sheet['B1'] = "Tweets_posted"

    for row, (Date, Tweets_posted) in enumerate(frequency_data.items(), start=2):
        frequency_sheet[f"A{row}"] = Date
        frequency_sheet[f"B{row}"] = Tweets_posted

    wb.save("C:\\Users\\Stephen Mac Donnacha\\PycharmProjects\\Final Year Project\\excel-worksheets\\fyp-data.xlsx")


def main():
    write_sentiment_data_to_excel()
    write_frequency_data_to_excel()


if __name__ == "__main__":
    main()
